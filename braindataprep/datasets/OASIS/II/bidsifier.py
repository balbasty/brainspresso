import tarfile
import nibabel as nib
import numpy as np
from tempfile import TemporaryDirectory
from logging import getLogger
from pathlib import Path, PosixPath
from typing import Literal, Iterable, Iterator

from braindataprep.utils.io import write_tsv
from braindataprep.utils.io import write_from_buffer
from braindataprep.utils.io import nibabel_convert
from braindataprep.utils.tabular import bidsify_tab
from braindataprep.utils.tabular import Status
from braindataprep.utils.vol import make_affine
from braindataprep.actions import IfExists
from braindataprep.actions import Action
from braindataprep.actions import CopyBytes
from braindataprep.actions import CopyJSON
from braindataprep.actions import WriteTSV
from braindataprep.actions import WrapAction

lg = getLogger(__name__)
try:
    import openpyxl
except ImportError:
    lg.error(
        'Cannot find `openpyxl`. Did you install with [oasis] flag? '
        'Try `pip install braindataprep[oasis]`.'
    )

KeyChoice = Literal["meta", "raw"]


class Bidsifier:
    """OASIS-II - bidsifying logic"""

    # ------------------------------------------------------------------
    #   Constants
    # ------------------------------------------------------------------

    # Folder containing template README/JSON/...
    TPLDIR = Path(__file__).parent / 'templates'

    AFFINE_RAW = make_affine(
        [256, 256, 128], [1.0, 1.0, 1.25], orient='ASL', center='x/2'
    )

    # ------------------------------------------------------------------
    #   Initialise
    # ------------------------------------------------------------------
    def __init__(
        self,
        root: Path,
        *,
        keys: Iterable[KeyChoice] = set(KeyChoice.__args__),
        parts: Iterable[int] = (1, 2),
        subs: Iterable[int] = tuple(),
        exclude_subs: Iterable[int] = tuple(),
        json: Literal["yes", "no", "only"] | bool = True,
        ifexists: IfExists.Choice = "skip",
    ):
        self.root: Path = Path(root)
        self.keys: set[KeyChoice] = set(keys)
        self.parts: set[int] = set(parts)
        self.subs: set[int] = set(subs)
        self.exclude_subs: set[int] = set(exclude_subs)
        self.json: Literal["yes", "no", "only"] = (
            "yes" if json is True else
            "no" if json is False else json
        )
        self.ifexists: IfExists.Choice = ifexists

    def init(self):
        """Prepare common stuff"""
        # Printer
        self.out = bidsify_tab()
        # Folder
        self.src = self.root / 'sourcedata'
        self.raw = self.root / 'rawdata'
        # Track errors
        self.nb_errors = 0
        self.nb_skipped = 0

    # ------------------------------------------------------------------
    #   Run all actions
    # ------------------------------------------------------------------
    def run(self):
        """Run all actions"""
        self.init()
        with self.out as self.out:
            self._run()

    def _run(self):
        """Must be run from inside the `out` context."""
        # Metadata
        if 'meta' in self.keys:
            self.nb_errors = self.nb_skipped = 0
            for status in self.make_meta():
                status.setdefault('modality', 'meta')
                self.out(status)

        # Raw and lightly processed data are stored in the same archive
        if 'raw' in self.keys:
            # prepare session data, which will be written in each subject
            self.make_sessions()
            for part in self.parts:
                self.nb_errors = self.nb_skipped = 0
                for status in self.make_raw(part):
                    status.setdefault('modality', f'raw ({part:02d})')
                    self.out(status)

    # ------------------------------------------------------------------
    #   Helpers
    # ------------------------------------------------------------------
    def fixstatus(self, status: Status, fname: str | Path) -> Iterator[Status]:
        status.setdefault('path', fname)
        yield status
        if status.get('status', '') == 'error':
            self.nb_errors += 1
            yield {'errors': self.nb_errors}
        elif status.get('status', '') == 'skipped':
            self.nb_skipped += 1
            yield {'skipped': self.nb_skipped}

    def tar2nii(
        self,
        tar: tarfile.TarFile,       # Opened TAR archive
        src: PosixPath,             # Member to unpack
        dst: Path,                  # Path to output nifti file
        affine: np.ndarray | None = None,
    ) -> Action:
        tarimg = str(src.with_suffix('.img'))
        tarhdr = str(src.with_suffix('.hdr'))

        def img2nii(niipath):
            with TemporaryDirectory() as tmp:
                tmp = Path(tmp)
                imgpath = tmp / 'tmp.img'
                hdrpath = tmp / 'tmp.hdr'
                # !!! we must unpack hdr before img
                # (that's how they are ordered in the stream)
                write_from_buffer(tar.extractfile(tarhdr), hdrpath)
                write_from_buffer(tar.extractfile(tarimg), imgpath)
                nibabel_convert(
                    imgpath,
                    niipath,
                    inp_format=nib.AnalyzeImage,
                    affine=affine,
                )

        return Action(Path(tar.name), dst, img2nii, input="path")

    # ------------------------------------------------------------------
    #   Write metadata files
    # ------------------------------------------------------------------
    def make_meta(self) -> Iterator[Status]:
        # Register future actions
        actions = {
            'README':
                CopyBytes(
                    self.TPLDIR / 'README',
                    self.root / 'README',
                ),
            'dataset_description.json':
                CopyJSON(
                    self.TPLDIR / 'dataset_description.json',
                    self.root / 'dataset_description.json',
                ),
            'participants.json':
                CopyJSON(
                    self.TPLDIR / 'participants.json',
                    self.root / 'participants.json',
                ),
            'participants.tsv':
                WrapAction(
                    self.src / 'oasis_longitudinal_demographics.xlsx',
                    self.root / 'participants.tsv',
                    self.make_participants,
                    mode="t", input="path",
                ),
            'sessions.json':
                CopyJSON(
                    self.TPLDIR / 'sessions.json',
                    self.root / 'sessions.json',
                ),
        }

        # Perform actions
        yield {'progress': 0}
        with IfExists(self.ifexists):
            for i, (fname, action) in enumerate(actions.items()):
                for status in action:
                    yield from self.fixstatus(status, fname)
                yield {'progress': 100*(i+1)/len(actions)}

        yield {'status': 'done', 'message': ''}

    # ------------------------------------------------------------------
    #   Write raw + processed data (except freesurfer)
    # ------------------------------------------------------------------
    def make_raw(self, part: int) -> Iterator[Status]:
        # Open tar file, then  delegate
        tarpath = self.src / f'OAS2_RAW_PART{part}.tar.gz'
        if not tarpath.exists():
            message = f'{tarpath.name} not found'
            lg.warning(message)
            yield {'status': 'error', 'message': message}
            return
        with tarfile.open(tarpath, 'r:gz') as tar:
            yield from self._make_raw(part, tar)

    def _make_raw(self, part: int, tar: tarfile.TarFile) -> Iterator[Status]:
        # 1. Find all subjects
        # 2. Write session files
        # 3. Iterate across files in the archive
        # 4. Convert each file to nifti
        #    (let's hope we extract hdr/img efficiently)
        subjects = self._raw_get_subjects(tar)

        # Make session.tsv
        for i, (id, _) in enumerate(subjects.items()):
            sub = self.raw / f'sub-{id:04d}'
            ses = sub / f'sub-{id:04d}_sessions.tsv'
            for status in WriteTSV(
                self.sessions_tables[id], ses,
                src=self.src / 'oasis_longitudinal_demographics.xlsx',
            ):
                yield from self.fixstatus(status, ses.name)

        # Make invididual scans
        # We iterate tar members rather than subject in the hope that
        # we can follow the gzip stream and be more efficient
        nscans = sum(map((lambda ses: sum(map(len, ses.values()))),
                         subjects.values()))
        nscan = 0
        yield {'progress': 0}
        with IfExists(self.ifexists):
            for member in tar:
                membername = PosixPath(member.name)
                if self._raw_skip_path(membername):
                    continue
                nscan += 1
                id, ses, run = self._raw_get_id(membername)
                for action in self._raw_get_actions(part, tar, id, ses, run):
                    for status in action:
                        yield from self.fixstatus(status, action.dst.name)
                yield {'progress': 100*nscan/nscans}
        yield {'status': 'done', 'message': ''}

    def _raw_get_id(self, path: PosixPath) -> tuple[int, int, int]:
        """Compute subject / session / run from path"""
        _, id, ses = path.parts[1].split('_')
        id, ses = int(id), int(ses[2:])
        run = int(path.name[4])
        return id, ses, run

    def _raw_skip_path(self, path: PosixPath):
        """Should we skip this member?"""

        def skip_subject(id: int) -> bool:
            return ((self.subs and id not in self.subs)
                    or id in self.exclude_subs)

        def skip_member(path: PosixPath) -> bool:
            if 'RAW' not in str(path):
                return True
            if path.suffix != '.img':
                return True
            if not path.name.startswith('mpr'):
                # Found a folder with a weird file 3906-3.nift.img
                # which seems to be a duplicate of mpr-1.nifti.img
                # Let's skip it
                return True
            return False

        if skip_member(path):
            return True
        id = self._raw_get_id(path)[0]
        if skip_subject(id):
            True
        return False

    def _raw_get_subjects(self, tar: tarfile.TarFile) -> dict[int, list[int]]:
        """Find all subject ids and runs contained in this archive"""
        subjects = {}
        for member in tar:
            path = PosixPath(member.name)
            if self._raw_skip_path(path):
                continue
            id, ses, run = self._raw_get_id(path)
            subjects.setdefault(id, {})
            subjects[id].setdefault(ses, [])
            subjects[id][ses].append(run)
        return subjects

    def _raw_get_actions(
        self,
        part: int,                       # part number
        tar: tarfile.TarFile,            # opened TAR archive
        id: int,                         # subject ID
        ses: int,                        # session ID
        run: int,                        # run ID
    ) -> Iterator[Action]:
        """Generate actions for a given subject"""
        json_path = self.TPLDIR / 'T1w.json'

        anat = self.raw / f'sub-{id:04d}' / f'ses-{ses}' / 'anat'
        base = anat / f'sub-{id:04d}_ses-{ses}_run-{run:d}_T1w'

        if self.json != 'no':
            yield CopyJSON(json_path, base.with_suffix('.json'))

        if self.json != 'only':
            member = PosixPath(
                f'OAS2_RAW_PART{part}/OAS2_{id:04d}_MR{ses}/RAW'
            )
            member = member / f'mpr-{run}.nifti.img'
            yield self.tar2nii(
                tar, member, base.with_suffix('.nii.gz'),
                affine=self.AFFINE_RAW
            )

    # ------------------------------------------------------------------
    #   Write participants.tsv
    # ------------------------------------------------------------------

    OASIS_HEADER = [
        'SUB_ID',
        'MRI_ID',
        'Group',
        'Visit',
        'Delay',
        'M/F',
        'Hand',
        'Age',
        'Educ',
        'SES',
        'MMSE',
        'CDR',
        'eTIV',
        'nWBV',
        'ASF',
    ]

    PARTICIPANTS_HEADER = [
        'participant_id',
        'sex',
        'handedness',
        'age',
    ]

    SESSION_HEADER = [
        'session_id',
        'delay',
        'pathology',
        'age',
        'educ',
        'ses',
        'mmse',
        'cdr',
        'etiv',
        'nwbv',
        'asf',
    ]

    PATHOLOGY_MAP = {
        "Nondemented": "N",
        "Demented": "D",
        "Converted": "C",
    }

    @classmethod
    def make_participants(cls, path_xlsx, path_tsv):

        def iter_rows():
            xlsx = openpyxl.load_workbook(path_xlsx, data_only=True)
            xlsx = xlsx[xlsx.sheetnames[0]]
            yield cls.PARTICIPANTS_HEADER
            for nrow in range(2, xlsx.max_row+1):
                id = xlsx[nrow][0].value
                visit = int(xlsx[nrow][3].value)
                if visit != 1:
                    continue
                sex = xlsx[nrow][5].value
                hand = xlsx[nrow][6].value
                age = xlsx[nrow][7].value
                id = int(id.split('_')[-1])
                yield [f'sub-{id:04d}', sex, hand, age]

        write_tsv(iter_rows(), path_tsv)

    def make_sessions(self):
        path_xlsx = self.src / 'oasis_longitudinal_demographics.xlsx'
        xlsx = openpyxl.load_workbook(path_xlsx, data_only=True)
        xlsx = xlsx[xlsx.sheetnames[0]]
        sessions_tables = {}
        for nrow in range(2, xlsx.max_row+1):
            id = int(xlsx[nrow][0].value.split('_')[-1])
            sessions_tables.setdefault(id, [self.SESSION_HEADER])
            sessions_tables[id].append([
                f'ses-{xlsx[nrow][3].value}',               # visit
                xlsx[nrow][4].value,                        # delay
                self.PATHOLOGY_MAP[xlsx[nrow][2].value],    # pathology
                xlsx[nrow][7].value,                        # age
                xlsx[nrow][8].value,                        # educ
                xlsx[nrow][9].value,                        # ses
                xlsx[nrow][10].value,                       # mmse
                xlsx[nrow][11].value,                       # cdr
                xlsx[nrow][12].value,                       # tiv
                xlsx[nrow][13].value,                       # nwbv
                xlsx[nrow][14].value,                       # asf
            ])
        self.sessions_tables = sessions_tables
