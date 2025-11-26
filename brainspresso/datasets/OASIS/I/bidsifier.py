import tarfile
import nibabel as nib
import numpy as np
import csv
from tempfile import TemporaryDirectory
from logging import getLogger
from pathlib import Path, PosixPath
from typing import Literal, Iterable, Iterator

from brainspresso.freesurfer import bidsify as fs
from brainspresso.utils.io import read_json
from brainspresso.utils.io import write_tsv
from brainspresso.utils.io import write_from_buffer
from brainspresso.utils.io import nibabel_convert
from brainspresso.utils.vol import make_affine
from brainspresso.utils.vol import relabel as vol_relabel
from brainspresso.utils.tabular import bidsify_tab
from brainspresso.utils.tabular import Status
from brainspresso.actions import IfExists
from brainspresso.actions import Action
from brainspresso.actions import CopyBytes
from brainspresso.actions import CopyJSON
from brainspresso.actions import WriteBytes
from brainspresso.actions import WriteJSON
from brainspresso.actions import WrapAction

lg = getLogger(__name__)
try:
    import openpyxl
except ImportError:
    lg.error(
        'Cannot find `openpyxl`. Did you install with [oasis] flag? '
        'Try `pip install brainspresso[oasis]`.'
    )

KeyChoice = Literal[
    "meta", "raw", "avg", "tal", "fsl", "fs", "fs-all"
]


class Bidsifier:
    """OASIS-I - bidsifying logic"""

    # ------------------------------------------------------------------
    #   Constants
    # ------------------------------------------------------------------

    # Folder containing template README/JSON/...
    TPLDIR = Path(__file__).parent / 'templates'

    # ------------------------------------------------------------------
    #   Initialise
    # ------------------------------------------------------------------
    def __init__(
        self,
        root: Path,
        *,
        keys: Iterable[KeyChoice] = set(KeyChoice.__args__),
        discs: Iterable[int] = range(1, 13),
        subs: Iterable[int] = tuple(),
        exclude_subs: Iterable[int] = tuple(),
        json: Literal["yes", "no", "only"] | bool = True,
        ifexists: IfExists.Choice = "skip",
    ):
        self.root: Path = Path(root)
        self.keys: set[KeyChoice] = set(keys)
        self.discs: set[int] = set(discs)
        self.subs: set[int] = set(subs)
        self.exclude_subs: set[int] = set(exclude_subs)
        self.json: Literal["yes", "no", "only"] = (
            "yes" if json is True else
            "no" if json is False else json
        )
        self.ifexists: IfExists.Choice = ifexists

    def init(self):
        """Prepare common stuff"""
        # Printer
        self.out = bidsify_tab()
        # Folder
        self.src = self.root / 'sourcedata'
        self.raw = self.root / 'rawdata'
        self.drv = self.root / 'derivatives'
        self.drvproc = self.drv / 'oasis-processed'
        self.drvfs = self.drv / 'oasis-freesurfer'
        # Track errors
        self.nb_errors = 0
        self.nb_skipped = 0

    # ------------------------------------------------------------------
    #   Run all actions
    # ------------------------------------------------------------------
    def run(self):
        """Run all actions"""
        self.init()
        with self.out as self.out:
            self._run()

    def _run(self):
        """Must be run from inside the `out` context."""
        # Metadata
        if 'meta' in self.keys:
            self.nb_errors = self.nb_skipped = 0
            for status in self.make_meta():
                status.setdefault('modality', 'meta')
                self.out(status)

        # Raw and lightly processed data are stored in the same archive
        if self.keys.intersection({'raw', 'avg', 'tal', 'fsl'}):
            for disc in self.discs:
                if not (1 <= disc <= 12):
                    continue
                self.nb_errors = self.nb_skipped = 0
                for status in self.make_raw(disc):
                    status.setdefault('modality', f'raw ({disc:02d})')
                    self.out(status)

        # Freesurfer outputs are stored in their own archive
        if self.keys.intersection({'fs', 'fs-all'}):
            for disc in self.discs:
                if not (1 <= disc <= 11):
                    # !!! no freesurfer disk 12
                    continue
                self.nb_errors = self.nb_skipped = 0
                for status in self.make_freesurfer(disc):
                    status.setdefault('modality', f'fs ({disc:02d})')
                    self.out(status)

    # ------------------------------------------------------------------
    #   Helpers
    # ------------------------------------------------------------------
    def fixstatus(self, status: Status, fname: str | Path) -> Iterator[Status]:
        status.setdefault('path', fname)
        yield status
        if status.get('status', '') == 'error':
            self.nb_errors += 1
            yield {'errors': self.nb_errors}
        elif status.get('status', '') == 'skipped':
            self.nb_skipped += 1
            yield {'skipped': self.nb_skipped}

    def tar2nii(
        self,
        tar: tarfile.TarFile,       # Opened TAR archive
        src: PosixPath,             # Member to unpack
        dst: Path,                  # Path to output nifti file
        affine: np.ndarray,         # 4x4 matrix
        relabel: bool = False,      # Assume FSL segmentation and relabel
    ) -> Action:
        tarimg = str(src.with_suffix('.img'))
        tarhdr = str(src.with_suffix('.hdr'))

        def img2nii(niipath):
            with TemporaryDirectory() as tmp:
                tmp = Path(tmp)
                imgpath = tmp / 'tmp.img'
                hdrpath = tmp / 'tmp.hdr'
                write_from_buffer(tar.extractfile(tarimg), imgpath)
                write_from_buffer(tar.extractfile(tarhdr), hdrpath)
                nibabel_convert(imgpath, niipath, affine=affine,
                                inp_format=nib.AnalyzeImage)
            if relabel:
                # Relabel FSL segmentation
                volf = nib.load(niipath)
                vold = np.asarray(volf.dataobj)
                vold = vol_relabel(vold, {1: 2, 2: 3, 3: 1})
                nib.save(type(volf)(vold, volf.affine, volf.header), dst)

        return Action(Path(tar.name), dst, img2nii, input="path")

    # ------------------------------------------------------------------
    #   Write metadata files
    # ------------------------------------------------------------------
    def make_meta(self) -> Iterator[Status]:
        # Register future actions
        actions = {
            'README':
                CopyBytes(
                    self.TPLDIR / 'README',
                    self.root / 'README',
                ),
            'dataset_description.json':
                CopyJSON(
                    self.TPLDIR / 'dataset_description.json',
                    self.root / 'dataset_description.json',
                ),
            'participants.json':
                CopyJSON(
                    self.TPLDIR / 'participants.json',
                    self.root / 'participants.json',
                ),
            'participants.tsv':
                WrapAction(
                    self.src / 'oasis_cross-sectional.xlsx',
                    self.root / 'participants.tsv',
                    self.make_participants,
                    mode="t", input="path",
                ),
        }

        # Register Freesurfer actions
        if self.keys.intersection({"fs", "fs-all"}):
            for action in fs.bidsify_toplevel(self.drvfs, (4, 0)):
                actions[str(action.dst.name)] = action

        # Perform actions
        yield {'progress': 0}
        with IfExists(self.ifexists):
            for i, (fname, action) in enumerate(actions.items()):
                for status in action:
                    yield from self.fixstatus(status, fname)
                yield {'progress': 100*(i+1)/len(actions)}

        yield {'status': 'done', 'message': ''}

    # ------------------------------------------------------------------
    #   Write raw + processed data (except freesurfer)
    # ------------------------------------------------------------------
    def make_raw(self, disc: int) -> Iterator[Status]:
        # Open tar file, then  delegate
        tarpath = self.src / f'oasis_cross-sectional_disc{disc}.tar.gz'
        if not tarpath.exists():
            message = f'oasis_cross-sectional_disc{disc}.tar.gz not found'
            lg.warning(message)
            yield {'status': 'error', 'message': message}
            return
        with tarfile.open(tarpath, 'r:gz') as tar:
            yield from self._make_raw(disc, tar)

    def _make_raw(self, disc: int, tar: tarfile.TarFile) -> Iterator[Status]:
        # 1. Find all subjects
        # 2. Iterate across subjects
        # 3. Iterate each subject's action
        # 4. Yield each action's statuses
        subjects = self._raw_get_subjects(tar)
        yield {'progress': 0}
        with IfExists(self.ifexists):
            for i, (id, runs) in enumerate(subjects.items()):
                for action in self._raw_get_actions(disc, tar, id, runs):
                    for status in action:
                        yield from self.fixstatus(status, action.dst.name)
                yield {'progress': 100*(i+1)/len(subjects)}
        yield {'status': 'done', 'message': ''}

    def _raw_get_subjects(self, tar: tarfile.TarFile) -> dict[int, list[int]]:
        """Find all subject ids and runs contained in this disc"""

        def skip_subject(id):
            return ((self.subs and id not in self.subs)
                    or id in self.exclude_subs)

        subjects = {}
        for path in tar.getnames():
            path = PosixPath(path)
            if 'RAW' not in str(path):
                continue
            if path.suffix != '.img':
                continue
            _, id, ses, run, _ = path.stem.split('_')
            id, ses, run = int(id), int(ses[2:]), int(run[4:])
            if skip_subject(id):
                continue
            if ses != 1:
                # skip repeats
                continue
            subjects.setdefault(id, [])
            subjects[id].append(run)
        return subjects

    def _raw_get_actions(
        self,
        disc: int,                  # disc number
        tar: tarfile.TarFile,       # opened TAR archive
        id: int,                    # Subject ID
        runs: list[int],            # Runs available in subject
    ) -> Iterator[Action]:
        """Generate actions for a given subject"""

        json_path = self.TPLDIR / 'T1w.json'
        json_base = read_json(json_path)

        # OASIS data has an ASL layout
        # https://brainder.org/2011/08/13/converting-oasis-brains-to-nifti/
        AFFINE_RAW = make_affine(
            [256, 256, 128], [1.0, 1.0, 1.25], orient='ASL', center='x/2'
        )
        AFFINE_AVG = make_affine(
            [256, 256, 160], [1.0, 1.0, 1.0], orient='ASL', center='x/2'
        )
        AFFINE_TAL = make_affine(
            [176, 208, 176], [1.0, 1.0, 1.0], orient='LAS', center='x/2'
        )

        # ----------------------------------------------------------
        #   Convert raw data
        #   (per-run scans only, the average is a derivative)
        # ----------------------------------------------------------
        def do_raw():
            anat = self.raw / f'sub-{id:04d}' / 'anat'
            for run in runs:
                base = f'sub-{id:04d}_run-{run:d}_T1w'

                if self.json != 'no':
                    name = base + '.json'
                    yield CopyJSON(json_path, anat / name)

                if self.json != 'only':
                    name = base + '.nii.gz'
                    member = PosixPath(f'disc{disc}/OAS1_{id:04d}_MR1/RAW')
                    member = member / f'OAS1_{id:04d}_MR1_mpr-{run:d}_anon'
                    yield self.tar2nii(tar, member, anat / name, AFFINE_RAW)

        if 'raw' in self.keys:
            yield from do_raw()

        # ----------------------------------------------------------
        #   Convert average scan
        #   (in derivative "oasis-processed")
        # ----------------------------------------------------------
        def do_avg():

            # which bias correction?
            member = (f'disc{disc}/OAS1_{id:04d}_MR1'
                      f'/PROCESSED/MPRAGE/SUBJ_111'
                      f'/OAS1_{id:04d}_MR1')
            member += '_mpr_{bias}_anon_sbj_111.img'
            if member.format(bias='n4') in tar.getnames():
                bias = 'n4'
            elif member.format(bias='n3') in tar.getnames():
                bias = 'n3'
            else:
                lg.error(f'Member not found: {member}')
                return
            member = PosixPath(member.format(bias=bias))

            srcbase = f'bids:raw:sub-{id:04d}/anat/sub-{id:04d}'
            json = {
                **json_base,
                "SkullStripped":
                    False,
                "BiasCorrected":
                    bias.upper(),
                "Resolution":
                    "Resampled and averaged across runs (1mm, isotropic)",
                "Sources": [
                    (srcbase + 'run-{run:d}_T1w.nii.gz').format(run=run)
                    for run in runs
                    ]
            }

            anat = self.drvproc / f'sub-{id:04d}' / 'anat'
            base = f'sub-{id:04d}_res-1mm_T1w'

            if self.json != 'no':
                name = base + '.json'
                yield WriteJSON(json, anat / name, src=json_path)

            if self.json != 'only':
                name = base + '.nii.gz'
                yield self.tar2nii(tar, member, anat / name, AFFINE_AVG)

        if 'avg' in self.keys:
            yield from do_avg()

        # ----------------------------------------------------------
        #   Convert talairach-transformed scan
        #   (in derivative "oasis-processed")
        # ----------------------------------------------------------
        def do_tal():

            # which bias correction?
            member = (f'disc{disc}/OAS1_{id:04d}_MR1'
                      f'/PROCESSED/MPRAGE/T88_111'
                      f'/OAS1_{id:04d}_MR1')
            member += '_mpr_{bias}_anon_111_t88_gfc'
            if member.format(bias='n4') in tar.getnames():
                bias = 'n4'
            elif member.format(bias='n3') in tar.getnames():
                bias = 'n3'
            else:
                lg.error(f'Member not found: {member}')
                return
            member = PosixPath(member.format(bias=bias))

            # json
            srcbase = f'bids:raw:sub-{id:04d}/anat/sub-{id:04d}'
            json = {
                **json_base,
                "SkullStripped":
                    False,
                "BiasCorrected":
                    bias.upper(),
                "Resolution":
                    "Resampled and averaged across runs (1mm, isotropic)",
                "Sources": [
                    (srcbase + 'run-{run:d}_T1w.nii.gz').format(run=run)
                    for run in runs
                    ]
            }

            # common
            anat = self.drvproc / f'sub-{id:04d}' / 'anat'

            # non-masked version
            base = f'sub-{id:04d}_space-Talairach_res-1mm_T1w'
            if self.json != 'no':
                name = base + '.json'
                yield WriteJSON(json, anat / name, src=json_path)

            if self.json != 'only':
                name = base + '.nii.gz'
                yield self.tar2nii(
                    tar, member, anat / name, AFFINE_TAL
                )

            # masked version
            flags = 'space-Talairach_res-1mm_desc-skullstripped'
            base = f'sub-{id:04d}_{flags}_T1w'
            if self.json != 'no':
                json['SkullStripped'] = True
                name = base + '.json'
                yield WriteJSON(json, anat / name, src=json_path)

            if self.json != 'only':
                name = base + '.nii.gz'
                member = PosixPath(str(member)[:-3] + 'masked_gfc')
                yield self.tar2nii(
                    tar, member, anat / name, AFFINE_TAL
                )

        if 'tal' in self.keys:
            yield from do_tal()

        # ----------------------------------------------------------
        #   Convert FSL segmentation
        #   (in derivative "oasis-processed")
        # ----------------------------------------------------------
        def do_fsl():

            # which bias correction?
            member = (f'disc{disc}/OAS1_{id:04d}_MR1/FSL_SEG'
                      f'/OAS1_{id:04d}_MR1')
            member += '_mpr_{bias}_anon_111_t88_masked_gfc_fseg.img'
            if member.format(bias='n4') in tar.getnames():
                bias = 'n4'
            elif member.format(bias='n3') in tar.getnames():
                bias = 'n3'
            else:
                lg.error(f'Member not found: {member}')
                return
            member = PosixPath(member.format(bias=bias))

            srcflags = 'space-Talairach_res-1mm_desc-skullstripped'
            json = {
                "Manual":
                    "False",
                "Resolution":
                    "In the space of the 1mm Talairach T1w scan "
                    "(1mm, isotropic)",
                "Sources": [
                    f"bids::sub-{id:04d}/anat/"
                    f"sub-{id:04d}_{srcflags}_T1w.nii.gz",
                ]
            }

            anat = self.drvproc / f'sub-{id:04d}' / 'anat'
            base = f'sub-{id:04d}_space-Talairach_res-1mm_dseg'

            if self.json != 'no':
                name = base + '.json'
                yield WriteJSON(json, anat / name, src=json_path)

            if self.json != 'only':
                name = base + '.nii.gz'
                yield self.tar2nii(
                    tar, member, anat / name, AFFINE_TAL, relabel=True
                )

        if 'fsl' in self.keys:
            yield from do_fsl()

    # ------------------------------------------------------------------
    #   Write freesurfer
    # ------------------------------------------------------------------
    def make_freesurfer(self, disc: int) -> Iterator[dict]:
        # Open tar archive then delegate
        tarpath = self.src / f'oasis_cs_freesurfer_disc{disc}.tar.gz'
        if not tarpath.exists():
            lg.warning(f'oasis_cross-sectional_disc{disc}.tar.gz not found')
            return
        with tarfile.open(tarpath, 'r:gz') as tar:
            yield from self._make_freesurfer(tar)

    def _make_freesurfer(self, tar: tarfile.TarFile) -> Iterator[dict]:
        # 1. Find all subjects
        # 2. Iterate across subjects
        # 3. Iterate each subject's action
        # 4. Yield each action's statuses
        subjects = self._fs_get_subjects(tar)
        yield {'progress': 0}
        with IfExists(self.ifexists):
            for i, (id, members) in enumerate(subjects.items()):
                for action in self._fs_get_actions(tar, id, members):
                    for status in action:
                        yield from self.fixstatus(status, action.dst.name)
                yield {'progress': 100*(i+1)/len(subjects)}
        yield {'status': 'done', 'message': ''}

    def _fs_get_subjects(self, tar: tarfile.TarFile):
        """Return dictionary subject id -> list of tar paths"""

        def skip_subject(id):
            return ((self.subs and id not in self.subs)
                    or id in self.exclude_subs)

        subjects = {}
        for path in tar.getnames():
            if 'fs-all' not in self.keys:
                if not path.endswith(fs.bidsifiable_outputs):
                    continue
            path = PosixPath(path)
            try:
                _, id, ses = path.parts[1].split('_')
            except Exception as e:
                lg.error(f'??? {str(path)}: {str(e)}')
                raise
            id, ses = int(id), int(ses[2:])
            if skip_subject(id):
                continue
            if ses != 1:
                # skip repeats
                continue
            subjects.setdefault(id, [])
            subjects[id].append(str(path))
        return subjects

    def _fs_get_actions(
            self,
            tar: tarfile.TarFile,
            id: int,
            members: list[str],
    ) -> Iterator[Action]:

        # Unpack raw freesurfer outputs
        # under "derivatives/oasis-freesurfer/sourcedata/sub-{04d}"
        for path in members:
            path = PosixPath(path)
            dst = self.drvfs / 'sourcedata' / f'sub-{id:04d}'
            dst = dst.joinpath(*path.parts[2:])
            yield WriteBytes(
                tar.extractfile(str(path)),
                dst,
                src=tar.name,
            )

        # Bidsify under "derivatives/oasis-freesurfer/sub-{04d}"
        src = self.drvfs / 'sourcedata' / f'sub-{id:04d}'
        dst = self.drvfs / f'sub-{id:04d}'
        srcbase = f'bids:raw:sub-{id:04d}/anat/sub-{id:04d}'
        sourcefiles = [
            srcbase + '_run-{:d}_T1w.nii.gz'.format(run)
            for run in (1, 2, 3, 4)
        ]
        yield from fs.bidsify(src, dst, sourcefiles, json=self.json)

    # ------------------------------------------------------------------
    #   Write participants.tsv
    # ------------------------------------------------------------------

    OASIS_HEADER = [
        'ID',
        'M/F',
        'Hand',
        'Age',
        'Educ',
        'SES',
        'MMSE',
        'CDR',
        'eTIV',
        'nWBV',
        'ASF',
        'Delay',
    ]

    PARTICIPANTS_HEADER = [
        'participant_id',
        'sex',
        'handedness'
        'age',
        'educ',
        'ses',
        'mmse',
        'cdr',
        'etiv',
        'nwbv',
        'asf',
    ]

    def make_participants(self, path_xlsx, path_tsv):
        # New version, since OASIS provides a Excel file

        xlsx = openpyxl.load_workbook(path_xlsx, data_only=True)
        xlsx = xlsx[xlsx.sheetnames[0]]

        def iter_rows():
            yield self.PARTICIPANTS_HEADER
            xlsx = openpyxl.load_workbook(path_xlsx, data_only=True)
            xlsx = xlsx[xlsx.sheetnames[0]]
            for nrow in range(2, xlsx.max_row+1):
                row = xlsx[nrow]
                id = row[0].value
                values = [row[i].value for i in range(1, 11)]
                values = ["n/a" if v in ('', 'N/A') else v for v in values]
                _, id, ses = id.split('_')
                id, ses = int(id), int(ses[2:])
                if ses == 2:
                    continue
                yield [f'sub-{id:04d}', *values]

        write_tsv(iter_rows(), path_tsv)

    def make_participants_csv(self, path_csv, path_tsv):
        # Old version, when OASIS was providing a CSV file

        def iter_rows():
            with open(path_csv, 'rt') as finp:
                yield self.PARTICIPANTS_HEADER
                reader = csv.reader(
                    finp, delimiter=',', quoting=csv.QUOTE_NONE
                )
                next(reader)  # skip header
                for row in reader:
                    id, *values = row
                    values = values[:-1]  # remove delay column
                    values = ["n/a" if v in ('', 'N/A') else v for v in values]
                    _, id, ses = id.split('_')
                    id, ses = int(id), int(ses[2:])
                    if ses == 2:
                        continue
                    yield [f'sub-{id:04d}', *values]

        write_tsv(iter_rows(), path_tsv)
